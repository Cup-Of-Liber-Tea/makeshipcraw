import json
import pandas as pd
import os
from datetime import datetime
import glob
import re

def load_json_files():
    """í˜„ì¬ í´ë”ì˜ Makeship ê´€ë ¨ JSON íŒŒì¼ë“¤ì„ ë¡œë“œ"""
    # 'makeship_all_products_YYYYMMDD_HHMMSS.json' íŒ¨í„´ê³¼ 'makeship_[ì¹´í…Œê³ ë¦¬]_[íƒ€ì„ìŠ¤íƒ¬í”„].json' íŒ¨í„´ì˜ íŒŒì¼ë“¤ì„ ëª¨ë‘ ì°¾ìŒ
    json_files = glob.glob('makeship_all_products_*.json') + \
                 glob.glob('makeship_*_*.json') # ëª¨ë“  makeship_ë¡œ ì‹œì‘í•˜ëŠ” json íŒŒì¼ í¬í•¨ (ì¹´í…Œê³ ë¦¬ë³„ íŒŒì¼ í¬í•¨)
    
    # ì¤‘ë³µ ì œê±° ë° ì •ë ¬
    json_files = sorted(list(set(json_files)))

    all_data = []
    
    print(f"ë°œê²¬ëœ Makeship JSON íŒŒì¼: {len(json_files)}ê°œ")
    
    for json_file in json_files:
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                
            # JSON êµ¬ì¡°ì— ë”°ë¼ ì œí’ˆ ëª©ë¡ ì¶”ì¶œ
            if 'ì œí’ˆ_ëª©ë¡' in data:
                products = data['ì œí’ˆ_ëª©ë¡']
                print(f"{json_file}: {len(products)}ê°œ ì œí’ˆ")
            else:
                # ë‹¨ì¼ ì œí’ˆ ë°ì´í„°ì¸ ê²½ìš° (ê³¼ê±° íŒŒì¼ í˜•ì‹ í˜¸í™˜)
                print(f"{json_file}: ë‹¨ì¼ ì œí’ˆ ë˜ëŠ” ì•Œ ìˆ˜ ì—†ëŠ” í˜•ì‹")
                products = [data]

            # ê° ì œí’ˆ ë°ì´í„°ì— ëŒ€í•´ í˜•ì‹ ë³€í™˜ ì ìš©
            for product in products:
                if 'í”„ë¡œì íŠ¸_ì¢…ë£Œì¼' in product: # í”„ë¡œì íŠ¸ ì¢…ë£Œì¼ ë‚ ì§œ í˜•ì‹ ë³€í™˜
                    product['í”„ë¡œì íŠ¸_ì¢…ë£Œì¼'] = normalize_date(product['í”„ë¡œì íŠ¸_ì¢…ë£Œì¼'])
                if 'ë°°ì†¡_ì‹œì‘ì¼' in product: # ë°°ì†¡ ì‹œì‘ì¼ ë‚ ì§œ í˜•ì‹ ë³€í™˜
                    product['ë°°ì†¡_ì‹œì‘ì¼'] = normalize_date(product['ë°°ì†¡_ì‹œì‘ì¼'])
                if 'íŒë§¤ëŸ‰' in product: # íŒë§¤ëŸ‰ ìˆ«ì í˜•ì‹ ë³€í™˜
                    product['íŒë§¤ëŸ‰'] = convert_to_numeric(product['íŒë§¤ëŸ‰'])
                if 'ë‹¬ì„±ë¥ ' in product: # ë‹¬ì„±ë¥  ìˆ«ì í˜•ì‹ ë³€í™˜
                    product['ë‹¬ì„±ë¥ '] = convert_to_numeric(product['ë‹¬ì„±ë¥ '])
                # íŒë§¤ëŸ‰ì´ 0ì´ë©´ ë§¤ì¶œë„ 0ìœ¼ë¡œ ì²˜ë¦¬
                if 'íŒë§¤ëŸ‰' in product and 'ë§¤ì¶œ' in product:
                    if product['íŒë§¤ëŸ‰'] == 0:
                        product['ë§¤ì¶œ'] = 0
            
            all_data.extend(products)
                
        except Exception as e:
            print(f"{json_file} ë¡œë“œ ì¤‘ ì˜¤ë¥˜: {e}")
    
    return all_data, json_files

def normalize_date(date_str):
    """
    ë‹¤ì–‘í•œ ë‚ ì§œ ë¬¸ìì—´ í˜•ì‹ì„ 'YYYY-MM-DD' í˜•ì‹ìœ¼ë¡œ ë³€í™˜í•©ë‹ˆë‹¤.
    - 'YYYY-MM-DD' (ì´ë¯¸ ì •ê·œí™”ëœ í˜•ì‹)
    - 'July 1, 5:00AM GMT+9 / Ships September 23, 2025'
    - 'September 17, 2022'
    - 'July 1, 2022'
    - 'í•´ë‹¹ ì—†ìŒ'
    ë“±ì„ ì²˜ë¦¬í•  ìˆ˜ ìˆë„ë¡ ê°œì„ í•©ë‹ˆë‹¤.
    """
    if not date_str or date_str in ['ì •ë³´ ì—†ìŒ', 'í•´ë‹¹ ì—†ìŒ', 'ìƒíƒœ í™•ì¸ ì¤‘ ì˜¤ë¥˜']:
        return date_str if date_str else 'ì •ë³´ ì—†ìŒ'

    # 1. ì´ë¯¸ YYYY-MM-DD í˜•ì‹ì¸ ê²½ìš° ê·¸ëŒ€ë¡œ ë°˜í™˜
    if isinstance(date_str, str) and re.match(r'^\d{4}-\d{2}-\d{2}$', date_str):
        return date_str

    # 2. ' / ' ê¸°ì¤€ìœ¼ë¡œ ë‚˜ëˆ„ì–´ í”„ë¡œì íŠ¸ ì¢…ë£Œì¼ê³¼ ë°°ì†¡ ì‹œì‘ì¼ ë¶„ë¦¬
    parts = date_str.split(' / ')
    
    # í”„ë¡œì íŠ¸ ì¢…ë£Œì¼ ì²˜ë¦¬ (ì²« ë²ˆì§¸ ë¶€ë¶„)
    project_end_date_part = parts[0].strip()
    try:
        # '5:00AM GMT+9'ì™€ ê°™ì€ ì‹œê°„/GMT ì •ë³´ ì œê±°
        project_end_date_clean = ' '.join(project_end_date_part.split(' ')[:3])
        # 'July 1, 2022' í˜•ì‹ íŒŒì‹±
        dt_object = datetime.strptime(project_end_date_clean.replace(',', ''), '%B %d %Y')
        return dt_object.strftime('%Y-%m-%d')
    except ValueError:
        pass # íŒŒì‹± ì‹¤íŒ¨ ì‹œ ë‹¤ìŒ í˜•ì‹ ì‹œë„

    # 3. 'Ships September 23, 2025' ê°™ì€ í˜•ì‹ ì²˜ë¦¬
    if 'Ships ' in date_str:
        ship_date_part = date_str.split('Ships ')[-1].strip()
        try:
            # 'September 23, 2025' í˜•ì‹ íŒŒì‹±
            dt_object = datetime.strptime(ship_date_part.replace(',', ''), '%B %d %Y')
            return dt_object.strftime('%Y-%m-%d')
        except ValueError:
            pass # íŒŒì‹± ì‹¤íŒ¨ ì‹œ ë‹¤ìŒ í˜•ì‹ ì‹œë„
    
    # 4. Fallback: ì¼ë°˜ì ì¸ ë‚ ì§œ í˜•ì‹ ì‹œë„
    try:
        dt_object = datetime.strptime(date_str.replace(',', ''), '%B %d %Y')
        return dt_object.strftime('%Y-%m-%d')
    except ValueError:
        # ë³€í™˜ ì‹¤íŒ¨ ì‹œ ì›ë³¸ ë°˜í™˜
        return date_str

def convert_to_numeric(value_str):
    """ë¬¸ìì—´ì—ì„œ ìˆ«ìë§Œ ì¶”ì¶œí•˜ì—¬ ì •ìˆ˜ ë˜ëŠ” ì‹¤ìˆ˜ë¡œ ë³€í™˜í•©ë‹ˆë‹¤."""
    # ì´ë¯¸ ìˆ«ìì¸ ê²½ìš° ê·¸ëŒ€ë¡œ ë°˜í™˜
    if isinstance(value_str, (int, float)):
        return value_str
    
    if not value_str or value_str == 'ì •ë³´ ì—†ìŒ':
        return 0
    
    # ë¬¸ìì—´ì´ ì•„ë‹Œ ê²½ìš° 0 ë°˜í™˜
    if not isinstance(value_str, str):
        return 0

    clean_value = value_str.replace(',', '').replace(' sold', '').strip()
    try:
        # ë‹¬ì„±ë¥  (%)ê°€ í¬í•¨ëœ ê²½ìš°
        if '%' in clean_value:
            return float(clean_value.replace('%', ''))
        else:
            # ì •ìˆ˜ ë˜ëŠ” ì‹¤ìˆ˜ë¡œ ë³€í™˜ ì‹œë„
            if '.' in clean_value:
                return float(clean_value)
            else:
                return int(clean_value)
    except ValueError:
        return 0

def remove_duplicates_by_url(products):
    """ì œí’ˆ URLë¡œ ì¤‘ë³µ ì œê±° (ìµœì‹  ë°ì´í„° ìœ ì§€)"""
    unique_products = {}
    
    for product in products:
        url = product.get('ì œí’ˆ_URL', '')
        if url:
            # ë™ì¼í•œ URLì´ ìˆìœ¼ë©´ ë®ì–´ì“°ê¸° (ìµœì‹  ë°ì´í„° ìœ ì§€)
            unique_products[url] = product
    
    return list(unique_products.values())

def create_excel_from_products(products, filename):
    """ì œí’ˆ ë°ì´í„°ë¥¼ ì—‘ì…€ íŒŒì¼ë¡œ ìƒì„±"""
    if not products:
        print(f"ë°ì´í„°ê°€ ì—†ì–´ì„œ {filename} íŒŒì¼ì„ ìƒì„±í•  ìˆ˜ ì—†ìŠµë‹ˆë‹¤.")
        return
    
    # ë°ì´í„°í”„ë ˆì„ ìƒì„±
    df = pd.DataFrame(products)
    
    # ì»¬ëŸ¼ ìˆœì„œ ì •ë¦¬
    column_order = [
        'ì œí’ˆ_URL',
        'ì§„í–‰_ì—¬ë¶€', 
        'ì œí’ˆêµ°',
        'ì œí’ˆëª…',
        'IPëª…',
        'IP_ì†Œê°œ_ë§í¬',
        'ì œí’ˆ_ê°€ê²©',
        'íŒë§¤ëŸ‰',
        'ë‹¬ì„±ë¥ ',
        'ë§¤ì¶œ',
        'í”„ë¡œì íŠ¸_ì¢…ë£Œì¼',
        'ë°°ì†¡_ì‹œì‘ì¼'
    ]
    
    # ì¡´ì¬í•˜ëŠ” ì»¬ëŸ¼ë§Œ ì„ íƒ
    available_columns = [col for col in column_order if col in df.columns]
    df = df[available_columns]

    # ìˆ«ìí˜• ì»¬ëŸ¼ íƒ€ì… í™•ì¸ ë° ë³€í™˜
    for col in ['íŒë§¤ëŸ‰', 'ë‹¬ì„±ë¥ ', 'ë§¤ì¶œ']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    
    # ì œí’ˆ ê°€ê²©ë„ ìˆ«ìë¡œ ë³€í™˜
    if 'ì œí’ˆ_ê°€ê²©' in df.columns:
        df['ì œí’ˆ_ê°€ê²©'] = pd.to_numeric(df['ì œí’ˆ_ê°€ê²©'], errors='coerce').fillna(0)

    # ë‚ ì§œ ì»¬ëŸ¼ì„ datetime ê°ì²´ë¡œ ë³€í™˜
    for col in ['í”„ë¡œì íŠ¸_ì¢…ë£Œì¼', 'ë°°ì†¡_ì‹œì‘ì¼']:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce', format='%Y-%m-%d')
    
    # ì—‘ì…€ íŒŒì¼ë¡œ ì €ì¥
    try:
        from openpyxl.styles import numbers
        
        with pd.ExcelWriter(filename, engine='openpyxl', datetime_format='yyyy-mm-dd') as writer:
            df.to_excel(writer, sheet_name='Products', index=False)
            
            # ì›Œí¬ì‹œíŠ¸ ì„œì‹ ì„¤ì •
            worksheet = writer.sheets['Products']
            
            # ì»¬ëŸ¼ë³„ í˜•ì‹ ì§€ì •
            for idx, col in enumerate(df.columns, start=1):
                col_letter = worksheet.cell(row=1, column=idx).column_letter
                
                # ìˆ«ì í˜•ì‹ ì§€ì •
                if col in ['íŒë§¤ëŸ‰', 'ë§¤ì¶œ']:
                    for row in range(2, len(df) + 2):
                        cell = worksheet.cell(row=row, column=idx)
                        cell.number_format = '#,##0'  # ì²œë‹¨ìœ„ êµ¬ë¶„ ì •ìˆ˜
                
                elif col == 'ë‹¬ì„±ë¥ ':
                    for row in range(2, len(df) + 2):
                        cell = worksheet.cell(row=row, column=idx)
                        cell.number_format = '0.0"%"'  # ì†Œìˆ˜ì  ì²«ì§¸ìë¦¬ + %
                
                elif col == 'ì œí’ˆ_ê°€ê²©':
                    for row in range(2, len(df) + 2):
                        cell = worksheet.cell(row=row, column=idx)
                        cell.number_format = '$#,##0.00'  # ë‹¬ëŸ¬ í˜•ì‹
                
                # ë‚ ì§œ í˜•ì‹ ì§€ì •
                elif col in ['í”„ë¡œì íŠ¸_ì¢…ë£Œì¼', 'ë°°ì†¡_ì‹œì‘ì¼']:
                    for row in range(2, len(df) + 2):
                        cell = worksheet.cell(row=row, column=idx)
                        cell.number_format = 'yyyy-mm-dd'
            
            # ì»¬ëŸ¼ ë„ˆë¹„ ìë™ ì¡°ì •
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        print(f"âœ… {filename} ìƒì„± ì™„ë£Œ ({len(df)}ê°œ ì œí’ˆ)")
        
    except Exception as e:
        print(f"âŒ {filename} ìƒì„± ì¤‘ ì˜¤ë¥˜: {e}")

def create_individual_excel_files(json_files):
    """ê° JSON íŒŒì¼ë³„ë¡œ ê°œë³„ ì—‘ì…€ íŒŒì¼ ìƒì„±"""
    print("\n=== ê°œë³„ ì—‘ì…€ íŒŒì¼ ìƒì„± ===")
    
    for json_file in json_files:
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # ì œí’ˆ ë°ì´í„° ì¶”ì¶œ
            if 'ì œí’ˆ_ëª©ë¡' in data:
                products = data['ì œí’ˆ_ëª©ë¡']
            else:
                products = [data]
            
            # ê° ì œí’ˆ ë°ì´í„°ì— ëŒ€í•´ í˜•ì‹ ë³€í™˜ ì ìš©
            for product in products:
                if 'í”„ë¡œì íŠ¸_ì¢…ë£Œì¼' in product:
                    product['í”„ë¡œì íŠ¸_ì¢…ë£Œì¼'] = normalize_date(product['í”„ë¡œì íŠ¸_ì¢…ë£Œì¼'])
                if 'ë°°ì†¡_ì‹œì‘ì¼' in product:
                    product['ë°°ì†¡_ì‹œì‘ì¼'] = normalize_date(product['ë°°ì†¡_ì‹œì‘ì¼'])
                if 'íŒë§¤ëŸ‰' in product:
                    product['íŒë§¤ëŸ‰'] = convert_to_numeric(product['íŒë§¤ëŸ‰'])
                if 'ë‹¬ì„±ë¥ ' in product:
                    product['ë‹¬ì„±ë¥ '] = convert_to_numeric(product['ë‹¬ì„±ë¥ '])
                # íŒë§¤ëŸ‰ì´ 0ì´ë©´ ë§¤ì¶œë„ 0ìœ¼ë¡œ ì²˜ë¦¬
                if 'íŒë§¤ëŸ‰' in product and 'ë§¤ì¶œ' in product:
                    if product['íŒë§¤ëŸ‰'] == 0:
                        product['ë§¤ì¶œ'] = 0
            
            # ì—‘ì…€ íŒŒì¼ëª… ìƒì„±
            excel_filename = json_file.replace('.json', '.xlsx')
            
            # ì—‘ì…€ íŒŒì¼ ìƒì„±
            create_excel_from_products(products, excel_filename)
            
        except Exception as e:
            print(f"âŒ {json_file} ì²˜ë¦¬ ì¤‘ ì˜¤ë¥˜: {e}")

def main():
    print("=== Makeship JSON to Excel ë³€í™˜ê¸° ===")
    print(f"ì‹¤í–‰ ì‹œê°„: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # 1. JSON íŒŒì¼ë“¤ ë¡œë“œ
    print("\n1. JSON íŒŒì¼ ë¡œë“œ ì¤‘...")
    all_products, json_files = load_json_files()
    
    if not all_products:
        print("âŒ ë³€í™˜í•  ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤.")
        return
    
    print(f"ì´ {len(all_products)}ê°œ ì œí’ˆ ë°ì´í„° ë¡œë“œ ì™„ë£Œ")
    
    # 2. ì¤‘ë³µ ì œê±°
    print("\n2. ì¤‘ë³µ ì œê±° ì¤‘...")
    unique_products = remove_duplicates_by_url(all_products)
    removed_count = len(all_products) - len(unique_products)
    print(f"ì¤‘ë³µ ì œê±° ì™„ë£Œ: {removed_count}ê°œ ì¤‘ë³µ ì œê±°, {len(unique_products)}ê°œ ê³ ìœ  ì œí’ˆ")
    
    # 3. í†µí•© ì—‘ì…€ íŒŒì¼ ìƒì„±
    print("\n3. í†µí•© ì—‘ì…€ íŒŒì¼ ìƒì„± ì¤‘...")
    integrated_filename = f"makeship_integrated_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    create_excel_from_products(unique_products, integrated_filename)
    
    # 4. ê°œë³„ ì—‘ì…€ íŒŒì¼ ìƒì„±
    print("\n4. ê°œë³„ ì—‘ì…€ íŒŒì¼ ìƒì„± ì¤‘...")
    create_individual_excel_files(json_files)
    
    # 5. ì™„ë£Œ ë³´ê³ 
    print(f"\n=== ë³€í™˜ ì™„ë£Œ ===")
    print(f"ğŸ“Š í†µí•© ì—‘ì…€: {integrated_filename}")
    print(f"ğŸ“ ê°œë³„ ì—‘ì…€: {len(json_files)}ê°œ íŒŒì¼")
    print(f"ğŸ”¢ ì´ ì œí’ˆ ìˆ˜: {len(unique_products)}ê°œ (ì¤‘ë³µ ì œê±° í›„)")

if __name__ == '__main__':
    main() 