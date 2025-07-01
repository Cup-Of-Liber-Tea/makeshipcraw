import re
from datetime import datetime
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("openpyxl이 설치되지 않았습니다. 'pip install openpyxl'로 설치해주세요.")

def convert_date_format(date_string):
    """
    다양한 날짜 형식을 '2025-07-01' 형태로 변환
    """
    if not date_string or date_string in ['해당 없음', '배송 시작일을 찾을 수 없습니다.', '배송 시작일을 찾을 수 없습니다. ']:
        return date_string
    
    # 이미 변환된 형식인지 확인
    if re.match(r'\d{4}-\d{2}-\d{2}', str(date_string)):
        return date_string
    
    try:
        # 1. "July 1, 5:00AM GMT+9" 형태 처리
        match = re.match(r'([A-Za-z]+) (\d{1,2}), (\d{1,2}:\d{2}[AP]M GMT\+9)', str(date_string))
        if match:
            month_name, day, time = match.groups()
            month_dict = {
                'January': 1, 'February': 2, 'March': 3, 'April': 4,
                'May': 5, 'June': 6, 'July': 7, 'August': 8,
                'September': 9, 'October': 10, 'November': 11, 'December': 12
            }
            if month_name in month_dict:
                return f"2025-{month_dict[month_name]:02d}-{int(day):02d}"
        
        # 2. "Ended: July 1, 2025" 형태 처리
        match = re.match(r'Ended: ([A-Za-z]+) (\d{1,2}), (\d{4})', str(date_string))
        if match:
            month_name, day, year = match.groups()
            month_dict = {
                'January': 1, 'February': 2, 'March': 3, 'April': 4,
                'May': 5, 'June': 6, 'July': 7, 'August': 8,
                'September': 9, 'October': 10, 'November': 11, 'December': 12
            }
            if month_name in month_dict:
                return f"{year}-{month_dict[month_name]:02d}-{int(day):02d}"
        
        # 3. "Ships July 1, 2025" 형태 처리
        match = re.match(r'Ships ([A-Za-z]+) (\d{1,2}), (\d{4})', str(date_string))
        if match:
            month_name, day, year = match.groups()
            month_dict = {
                'January': 1, 'February': 2, 'March': 3, 'April': 4,
                'May': 5, 'June': 6, 'July': 7, 'August': 8,
                'September': 9, 'October': 10, 'November': 11, 'December': 12
            }
            if month_name in month_dict:
                return f"{year}-{month_dict[month_name]:02d}-{int(day):02d}"
        
        # 4. "Estimated to Ship: July 2025" 형태 처리
        match = re.match(r'Estimated to Ship: ([A-Za-z]+) (\d{4})', str(date_string))
        if match:
            month_name, year = match.groups()
            month_dict = {
                'January': 1, 'February': 2, 'March': 3, 'April': 4,
                'May': 5, 'June': 6, 'July': 7, 'August': 8,
                'September': 9, 'October': 10, 'November': 11, 'December': 12
            }
            if month_name in month_dict:
                return f"{year}-{month_dict[month_name]:02d}-01"
        
        # 5. 이모지가 포함된 형태 처리 (🚨, ⏰)
        # "July 1, 1:00AM GMT+9 🚨" 또는 "July 1, 1:00AM GMT+9 ⏰" 형태
        match = re.match(r'([A-Za-z]+) (\d{1,2}), (\d{1,2}:\d{2}[AP]M GMT\+9) [🚨⏰]', str(date_string))
        if match:
            month_name, day, time = match.groups()
            month_dict = {
                'January': 1, 'February': 2, 'March': 3, 'April': 4,
                'May': 5, 'June': 6, 'July': 7, 'August': 8,
                'September': 9, 'October': 10, 'November': 11, 'December': 12
            }
            if month_name in month_dict:
                return f"2025-{month_dict[month_name]:02d}-{int(day):02d}"
        
        # 6. "June 30, 1:00AM GMT+9 🚨" 형태 (년도 없음)
        match = re.match(r'([A-Za-z]+) (\d{1,2}), (\d{1,2}:\d{2}[AP]M GMT\+9) [🚨⏰]', str(date_string))
        if match:
            month_name, day, time = match.groups()
            month_dict = {
                'January': 1, 'February': 2, 'March': 3, 'April': 4,
                'May': 5, 'June': 6, 'July': 7, 'August': 8,
                'September': 9, 'October': 10, 'November': 11, 'December': 12
            }
            if month_name in month_dict:
                return f"2025-{month_dict[month_name]:02d}-{int(day):02d}"
        
        # 원래 문자열 반환 (변환할 수 없는 경우)
        return str(date_string)
        
    except Exception as e:
        print(f"날짜 변환 오류: {date_string} -> {e}")
        return str(date_string)

def convert_excel_dates():
    """
    Excel 파일의 I열과 J열 날짜 형식을 변환
    """
    file_path = r'e:\dev\crow\makeship\최종합본_.xlsx'
    
    try:
        # openpyxl로 워크북 로드
        wb = load_workbook(file_path)
        ws = wb.active
        
        # I열과 J열 데이터 변환
        print("날짜 형식 변환 시작...")
        
        # 변환된 데이터 카운트
        converted_count = 0
        
        # 2행부터 마지막 행까지 처리 (1행은 헤더)
        for row in range(2, ws.max_row + 1):
            # I열 (프로젝트 종료일) 변환
            i_cell = ws[f'I{row}']
            if i_cell.value:
                original_value = str(i_cell.value)
                converted_value = convert_date_format(original_value)
                if converted_value != original_value:
                    i_cell.value = converted_value
                    converted_count += 1
                    print(f"I{row}: {original_value} -> {converted_value}")
            
            # J열 (배송 시작일) 변환
            j_cell = ws[f'J{row}']
            if j_cell.value:
                original_value = str(j_cell.value)
                converted_value = convert_date_format(original_value)
                if converted_value != original_value:
                    j_cell.value = converted_value
                    converted_count += 1
                    print(f"J{row}: {original_value} -> {converted_value}")
        
        # 변환된 파일 저장
        output_file = r'e:\dev\crow\makeship\최종합본_날짜변환_ISO.xlsx'
        wb.save(output_file)
        
        print(f"\n변환 완료!")
        print(f"총 {converted_count}개 항목이 변환되었습니다.")
        print(f"변환된 파일: {output_file}")
        
    except Exception as e:
        print(f"파일 처리 오류: {e}")

if __name__ == "__main__":
    # 테스트용 샘플 데이터
    test_dates = [
        "July 1, 5:00AM GMT+9",
        "Ended: July 2, 2024",
        "Ships September 23, 2025",
        "Estimated to Ship: July 2025",
        "June 30, 1:00AM GMT+9 🚨",
        "July 7, 9:00PM GMT+9 ⏰",
        "해당 없음",
        "배송 시작일을 찾을 수 없습니다."
    ]
    
    print("=== 날짜 변환 테스트 ===")
    for date in test_dates:
        converted = convert_date_format(date)
        print(f"{date} -> {converted}")
    
    print("\n=== Excel 파일 변환 실행 ===")
    convert_excel_dates()
